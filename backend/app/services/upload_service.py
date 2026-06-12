"""Upload service — store files, parse structure, profile columns, register CIP graph.

Phase 2 additions:
  - SHA-256 file hash (immutability + duplicate detection)
  - CSV and XLSX parsing into RawDataset + RawField records
  - Column profiling: inferred type, sample values, missing/distinct counts, min/max
  - CIP context graph registration: UploadedFile, RawDataset, RawField nodes + edges
"""

from __future__ import annotations

import csv
import hashlib
import io
import json
import re
import uuid
import zipfile
from pathlib import Path
from uuid import UUID

from fastapi import HTTPException, UploadFile, status
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.config import get_settings
from app.core.permissions import Permission, check_permission
from app.core.phi_masking import mask_sample_values
from app.models.audit import AuditAction
from app.models.graph import GraphEdgeType, GraphNodeType
from app.models.intelligence import DataLineageType
from app.models.data_source import DataSourceType
from app.models.upload import UploadedFile
from app.models.user import User
from app.services.data_cut_service import DataCutContext
from app.repositories.raw_data_repository import (
    RawDatasetRepository,
    RawFieldRepository,
)
from app.repositories.study_repository import StudyRepository
from app.repositories.upload_repository import UploadRepository
from app.services.audit_service import AuditService
from app.services.context_graph_service import ContextGraphService
from app.services.intelligence_service import DataLineageService
from app.services.storage_service import StorageService, get_storage_service

_PARSEABLE_MIME_TYPES = {
    "text/csv",
    "application/vnd.ms-excel",
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
}
_EXTENSION_MIME_TYPES = {
    ".csv": {
        "canonical": "text/csv",
        "declared": {"text/csv", "application/vnd.ms-excel"},
    },
    ".xls": {
        "canonical": "application/vnd.ms-excel",
        "declared": {"application/vnd.ms-excel"},
    },
    ".xlsx": {
        "canonical": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "declared": {
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        },
    },
    ".pdf": {
        "canonical": "application/pdf",
        "declared": {"application/pdf"},
    },
    ".txt": {
        "canonical": "text/plain",
        "declared": {"text/plain"},
    },
}
_OLE2_SIGNATURE = bytes.fromhex("D0CF11E0A1B11AE1")
_MAX_FILE_SIZE = 50 * 1024 * 1024  # 50 MB
_MAX_PROFILE_ROWS = 10_000  # cap profiling scan to avoid OOM


class UploadService:
    """Handles file storage, parsing, profiling, and CIP graph registration."""

    def __init__(
        self,
        db: AsyncSession,
        storage: StorageService | None = None,
    ) -> None:
        self._db = db
        self._repo = UploadRepository(db)
        self._study_repo = StudyRepository(db)
        self._audit = AuditService(db)
        self._ds_repo = RawDatasetRepository(db)
        self._field_repo = RawFieldRepository(db)
        self._graph = ContextGraphService(db)
        self._lineage = DataLineageService(db)
        self._settings = get_settings()
        self._storage = storage or get_storage_service()

    async def upload_file(
        self,
        study_id: UUID,
        actor: User,
        file: UploadFile,
        description: str | None = None,
        data_source_type: DataSourceType = DataSourceType.LIVE_FINAL,
        data_cut_label: str | None = None,
        data_cut_date=None,
        notes: str | None = None,
        ip_address: str | None = None,
        user_agent: str | None = None,
    ) -> UploadedFile:
        """Store a file, parse/profile its columns, and register in the CIP graph."""
        check_permission(actor, Permission.ARTIFACT_CREATE)
        await self._study_repo.get(study_id, actor.organization_id)

        content = await file.read()
        if len(content) > _MAX_FILE_SIZE:
            raise HTTPException(
                status_code=status.HTTP_413_REQUEST_ENTITY_TOO_LARGE,
                detail={
                    "code": "FILE_TOO_LARGE",
                    "message": "File exceeds 50 MB limit.",
                },
            )

        mime_type = self._validate_upload_type(
            content=content,
            filename=file.filename or "",
            declared_mime_type=file.content_type,
        )

        file_hash = hashlib.sha256(content).hexdigest()
        stored_filename = f"{uuid.uuid4()}_{file.filename or 'upload'}"
        storage_key = self._storage.put(
            f"studies/{study_id}/uploads/{stored_filename}",
            content,
            organization_id=actor.organization_id,
        )

        extracted_metadata = self._extract_legacy_metadata(content, mime_type)

        filename = file.filename or ""
        if (
            data_source_type == DataSourceType.LIVE_FINAL
            and self._detect_synthetic_upload(content, mime_type, filename)
        ):
            data_source_type = DataSourceType.SYNTHETIC

        if data_source_type == DataSourceType.LIVE_INTERIM and not data_cut_label:
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail={
                    "code": "DATA_CUT_LABEL_REQUIRED",
                    "message": "Interim live uploads require a data cut label.",
                },
            )

        is_synthetic = data_source_type == DataSourceType.SYNTHETIC
        if is_synthetic:
            synthetic_version = (
                await self._count_synthetic_uploads(study_id, actor.organization_id) + 1
            )
            default_label = (
                data_cut_label or f"Synthetic Data Version {synthetic_version}"
            )
        else:
            default_label = data_cut_label or (
                "Final Data Cut"
                if data_source_type == DataSourceType.LIVE_FINAL
                else "Interim Data Cut"
            )

        record = await self._repo.create(
            organization_id=actor.organization_id,
            study_id=study_id,
            uploaded_by_id=actor.id,
            original_filename=file.filename or stored_filename,
            stored_filename=stored_filename,
            file_path=storage_key,
            file_size_bytes=len(content),
            mime_type=mime_type,
            description=description or notes,
            extracted_metadata=extracted_metadata,
            file_hash=file_hash,
            upload_status="UPLOADED",
            data_source_type=data_source_type,
            data_cut_label=default_label,
            data_cut_date=data_cut_date,
            is_synthetic=is_synthetic,
        )
        record.data_cut_id = record.id
        await self._repo.update(record)

        await self._audit.log(
            action=AuditAction.DATA_FILE_UPLOADED,
            resource_type="uploaded_file",
            organization_id=actor.organization_id,
            actor_user_id=actor.id,
            resource_id=record.id,
            after_state={
                "study_id": str(study_id),
                "filename": file.filename,
                "mime_type": mime_type,
                "size_bytes": len(content),
                "file_hash": file_hash,
                "data_source_type": data_source_type.value,
                "data_cut_label": default_label,
                "is_synthetic": is_synthetic,
            },
            ip_address=ip_address,
            user_agent=user_agent,
        )

        # Parse and profile if it's a structured file
        if mime_type in _PARSEABLE_MIME_TYPES:
            await self._parse_and_profile(
                record=record,
                content=content,
                mime_type=mime_type,
                filename=file.filename or "",
                actor=actor,
                study_id=study_id,
            )

        await self._db.commit()
        return record

    async def _parse_and_profile(
        self,
        record: UploadedFile,
        content: bytes,
        mime_type: str,
        filename: str,
        actor: User,
        study_id: UUID,
    ) -> None:
        """Parse file into sheets/columns, profile each column, register graph."""
        try:
            sheets = self._parse_file(content, mime_type, filename)
            record.upload_status = "PARSED"

            # Register UploadedFile node in the CIP graph
            if (
                record.is_synthetic
                or record.data_source_type == DataSourceType.SYNTHETIC
            ):
                ver_match = re.search(r"(\d+)$", record.data_cut_label or "")
                version_number = int(ver_match.group(1)) if ver_match else 1
                data_cut = DataCutContext.for_synthetic_upload(
                    study_id=study_id,
                    created_by=actor.id,
                    upload_id=record.id,
                    version_number=version_number,
                    data_cut_label=record.data_cut_label,
                    created_at=record.created_at,
                )
            else:
                data_cut = DataCutContext.for_live_upload(
                    study_id=study_id,
                    created_by=actor.id,
                    upload_id=record.id,
                    data_source_type=record.data_source_type,
                    data_cut_label=record.data_cut_label or "Data Cut",
                    data_cut_date=record.data_cut_date,
                    notes=record.description,
                    created_at=record.created_at,
                )
            file_node, _ = await self._graph.register_domain_record(
                organization_id=actor.organization_id,
                node_type=GraphNodeType.UPLOADED_FILE,
                external_id=record.id,
                external_type="uploaded_file",
                label=f"{record.original_filename} — {data_cut.data_cut_label}",
                study_id=study_id,
                properties=data_cut.to_dict(),
                actor=actor,
            )

            # Get the Study graph node (for the belongs_to edge)
            study_node, _ = await self._graph.register_domain_record(
                organization_id=actor.organization_id,
                node_type=GraphNodeType.STUDY,
                external_id=study_id,
                external_type="study",
                label=str(study_id),
                study_id=study_id,
                actor=actor,
            )

            await self._graph.create_relationship(
                organization_id=actor.organization_id,
                source_node_id=file_node.id,
                target_node_id=study_node.id,
                edge_type=GraphEdgeType.PART_OF,
                study_id=study_id,
                actor=actor,
            )

            uploader_node, _ = await self._graph.register_domain_record(
                organization_id=actor.organization_id,
                node_type=GraphNodeType.USER,
                external_id=actor.id,
                external_type="user",
                label=actor.email,
                study_id=study_id,
                actor=actor,
            )
            await self._graph.create_relationship(
                organization_id=actor.organization_id,
                source_node_id=file_node.id,
                target_node_id=uploader_node.id,
                edge_type=GraphEdgeType.CREATED_BY,
                study_id=study_id,
                actor=actor,
            )

            total_fields = 0
            for sheet in sheets:
                ds = await self._ds_repo.create(
                    organization_id=actor.organization_id,
                    study_id=study_id,
                    uploaded_file_id=record.id,
                    dataset_name=sheet["name"],
                    row_count=sheet["row_count"],
                    column_count=len(sheet["columns"]),
                    parse_status="PARSED",
                    data_source_type=record.data_source_type,
                    data_cut_label=record.data_cut_label,
                    data_cut_date=record.data_cut_date,
                    is_synthetic=record.is_synthetic,
                    data_cut_id=record.data_cut_id,
                    source_upload_id=record.id,
                )

                ds_node, _ = await self._graph.register_domain_record(
                    organization_id=actor.organization_id,
                    node_type=GraphNodeType.RAW_DATASET,
                    external_id=ds.id,
                    external_type="raw_dataset",
                    label=f"{record.original_filename} / {sheet['name']}",
                    study_id=study_id,
                    actor=actor,
                )

                await self._graph.create_relationship(
                    organization_id=actor.organization_id,
                    source_node_id=ds_node.id,
                    target_node_id=file_node.id,
                    edge_type=GraphEdgeType.DERIVED_FROM,
                    study_id=study_id,
                    actor=actor,
                )

                for col in sheet["columns"]:
                    field = await self._field_repo.create(
                        organization_id=actor.organization_id,
                        study_id=study_id,
                        raw_dataset_id=ds.id,
                        column_name=col["name"],
                        column_index=col["index"],
                        inferred_type=col["inferred_type"],
                        sample_values=col["sample_values"],
                        missing_count=col["missing_count"],
                        distinct_count=col["distinct_count"],
                        min_value=col.get("min_value"),
                        max_value=col.get("max_value"),
                    )

                    field_node, _ = await self._graph.register_domain_record(
                        organization_id=actor.organization_id,
                        node_type=GraphNodeType.RAW_DATA_FIELD,
                        external_id=field.id,
                        external_type="raw_field",
                        label=col["name"],
                        study_id=study_id,
                        actor=actor,
                    )

                    await self._graph.create_relationship(
                        organization_id=actor.organization_id,
                        source_node_id=field_node.id,
                        target_node_id=ds_node.id,
                        edge_type=GraphEdgeType.PART_OF,
                        study_id=study_id,
                        actor=actor,
                    )

                    await self._lineage.record_field_lineage(
                        organization_id=actor.organization_id,
                        lineage_type=DataLineageType.DERIVED,
                        source_type="uploaded_file",
                        source_id=record.id,
                        source_domain=sheet["name"],
                        target_type="raw_field",
                        target_id=field.id,
                        target_field=col["name"],
                        transformation_logic="CSV/XLSX column extraction and profiling",
                        study_id=study_id,
                        created_by=actor,
                        source_graph_node_id=file_node.id,
                        target_graph_node_id=field_node.id,
                    )
                    total_fields += 1

            await self._audit.log(
                action=AuditAction.DATA_FILE_PARSED,
                resource_type="uploaded_file",
                organization_id=actor.organization_id,
                actor_user_id=actor.id,
                resource_id=record.id,
                after_state={
                    "study_id": str(study_id),
                    "filename": record.original_filename,
                    "upload_status": record.upload_status,
                    "dataset_count": len(sheets),
                    "field_count": total_fields,
                },
            )

        except Exception as exc:
            record.upload_status = "FAILED"
            record.extracted_metadata = {
                **record.extracted_metadata,
                "parse_error": str(exc),
            }

    async def list_for_study(
        self,
        study_id: UUID,
        organization_id: UUID,
        page: int = 1,
        page_size: int = 25,
    ) -> tuple[list[UploadedFile], int]:
        """Return paginated file list for a study."""
        offset = (page - 1) * page_size
        return await self._repo.list_for_study(
            study_id=study_id,
            organization_id=organization_id,
            limit=page_size,
            offset=offset,
        )

    async def get_file(self, file_id: UUID, organization_id: UUID) -> UploadedFile:
        """Return a single file record; raise 404 if not found."""
        record = await self._repo.get_by_id(file_id, organization_id)
        if record is None:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail={"code": "NOT_FOUND", "message": "Upload not found."},
            )
        return record

    # ------------------------------------------------------------------
    # Parsing helpers
    # ------------------------------------------------------------------

    @staticmethod
    def _validate_upload_type(
        *,
        content: bytes,
        filename: str,
        declared_mime_type: str | None,
    ) -> str:
        """Require file extension, declared MIME, and detected content to agree."""
        extension = Path(filename).suffix.lower()
        spec = _EXTENSION_MIME_TYPES.get(extension)
        if spec is None:
            raise HTTPException(
                status_code=status.HTTP_415_UNSUPPORTED_MEDIA_TYPE,
                detail={
                    "code": "UNSUPPORTED_MEDIA_TYPE",
                    "message": "Unsupported file extension. Supported: CSV, XLS, XLSX, PDF, TXT.",
                },
            )

        declared = (
            (declared_mime_type or "application/octet-stream")
            .split(";", 1)[0]
            .strip()
            .lower()
        )
        if declared not in spec["declared"] or not UploadService._content_matches_type(
            content, extension
        ):
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail={
                    "code": "FILE_TYPE_MISMATCH",
                    "message": (
                        "File extension, declared Content-Type, and detected content "
                        "must agree."
                    ),
                },
            )
        return str(spec["canonical"])

    @staticmethod
    def _content_matches_type(content: bytes, extension: str) -> bool:
        if extension in {".csv", ".txt"}:
            return UploadService._is_text_content(content)
        if extension == ".pdf":
            return content[:1024].lstrip().startswith(b"%PDF-")
        if extension == ".xls":
            return content.startswith(_OLE2_SIGNATURE)
        if extension == ".xlsx":
            try:
                with zipfile.ZipFile(io.BytesIO(content)) as archive:
                    names = set(archive.namelist())
                    return {
                        "[Content_Types].xml",
                        "xl/workbook.xml",
                    }.issubset(names)
            except (OSError, zipfile.BadZipFile):
                return False
        return False

    @staticmethod
    def _is_text_content(content: bytes) -> bool:
        try:
            text = content.decode("utf-8-sig")
        except UnicodeDecodeError:
            return False
        return all(char in "\t\n\r" or char.isprintable() for char in text)

    @staticmethod
    def _parse_file(content: bytes, mime_type: str, filename: str) -> list[dict]:
        """Return a list of sheet dicts, each with name, row_count, and columns."""
        if mime_type == "text/csv":
            return UploadService._parse_csv(content, filename)
        if mime_type in (
            "application/vnd.ms-excel",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        ):
            return UploadService._parse_xlsx(content)
        return []

    @staticmethod
    def _parse_csv(content: bytes, filename: str) -> list[dict]:
        text = content.decode("utf-8", errors="replace")
        reader = csv.reader(io.StringIO(text))
        rows = list(reader)
        if not rows:
            return []
        headers = rows[0]
        data_rows = rows[1:_MAX_PROFILE_ROWS]
        columns = []
        for col_idx, header in enumerate(headers):
            col_vals = []
            for row in data_rows:
                col_vals.append(row[col_idx] if col_idx < len(row) else None)
            col_name = header or f"col_{col_idx}"
            columns.append(
                {
                    "name": col_name,
                    "index": col_idx,
                    **UploadService._profile_column(col_name, col_vals),
                }
            )
        return [{"name": filename, "row_count": len(data_rows), "columns": columns}]

    @staticmethod
    def _sanitize_sheet_name(name: str) -> str:
        """Normalize Excel sheet names for stable dataset identifiers."""
        cleaned = "".join(
            ch if ch.isalnum() or ch in " _-" else "_" for ch in name.strip()
        )
        cleaned = cleaned.strip(" _") or "Sheet"
        return cleaned[:100]

    @staticmethod
    def _parse_xlsx(content: bytes) -> list[dict]:
        import openpyxl

        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True, read_only=True)
        sheets = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            safe_name = UploadService._sanitize_sheet_name(sheet_name)
            rows = list(ws.values)
            if not rows:
                continue
            headers = [
                str(h) if h is not None else f"col_{i}" for i, h in enumerate(rows[0])
            ]
            data_rows = rows[1:_MAX_PROFILE_ROWS]
            columns = []
            for col_idx, header in enumerate(headers):
                col_vals = [
                    str(row[col_idx])
                    if (col_idx < len(row) and row[col_idx] is not None)
                    else None
                    for row in data_rows
                ]
                columns.append(
                    {
                        "name": header,
                        "index": col_idx,
                        **UploadService._profile_column(header, col_vals),
                    }
                )
            sheets.append(
                {
                    "name": safe_name,
                    "source_sheet_name": sheet_name,
                    "row_count": len(data_rows),
                    "columns": columns,
                }
            )
        wb.close()
        return sheets

    @staticmethod
    def resolve_storage_path(file_path: str, storage_root: str) -> Path:
        """Resolve an absolute or storage-relative upload path."""
        path = Path(file_path)
        if path.is_absolute():
            return path
        return Path(storage_root) / file_path

    @staticmethod
    def reconstruct_rows_from_fields(
        fields: list,
        row_count: int,
        *,
        max_rows: int = 500,
    ) -> list[dict]:
        """Rebuild row dicts from profiled column sample values when the file is missing."""
        if not fields:
            return []

        max_samples = max(
            (len(getattr(f, "sample_values", []) or []) for f in fields), default=0
        )
        effective_rows = min(max(row_count, max_samples), max_rows)
        if effective_rows <= 0:
            return []

        ordered = sorted(fields, key=lambda f: getattr(f, "column_index", 0))
        rows_out: list[dict] = []
        for row_idx in range(effective_rows):
            record: dict[str, str | None] = {}
            for field in ordered:
                samples = getattr(field, "sample_values", []) or []
                if not samples:
                    continue
                record[field.column_name] = str(samples[row_idx % len(samples)])
            if record:
                rows_out.append(record)
        return rows_out

    @staticmethod
    def read_json_edc_rows(
        content: bytes,
        dataset_name: str,
        *,
        max_rows: int = 500,
    ) -> list[dict]:
        """Read rows from a synthetic RAW_CLINICAL_DATA JSON export."""
        try:
            payload = json.loads(content.decode("utf-8"))
        except (json.JSONDecodeError, UnicodeDecodeError):
            return []

        datasets = payload.get("datasets", {})
        if not isinstance(datasets, dict):
            return []

        form_id = dataset_name.split(" — ", 1)[0].strip()
        sheet = None
        for candidate in datasets.values():
            if not isinstance(candidate, dict):
                continue
            if (
                candidate.get("form_id") == form_id
                or candidate.get("form_name") == dataset_name
            ):
                sheet = candidate
                break
        if sheet is None and datasets:
            sheet = next(iter(datasets.values()))

        if not isinstance(sheet, dict):
            return []

        columns = sheet.get("columns", [])
        sample_rows = sheet.get("sample_rows", [])
        if not isinstance(columns, list) or not isinstance(sample_rows, list):
            return []

        rows_out: list[dict] = []
        for row in sample_rows[:max_rows]:
            if not isinstance(row, dict):
                continue
            record = {col: row.get(col) for col in columns if col in row}
            if record:
                rows_out.append(record)
        return rows_out

    @staticmethod
    def read_tabular_rows(
        *,
        file_path: str,
        mime_type: str,
        filename: str,
        dataset_name: str,
        max_rows: int = 500,
        storage_root: str | None = None,
        storage: StorageService | None = None,
    ) -> list[dict]:
        """Read raw row dicts from a stored CSV/XLSX/JSON file for SDTM derivation."""
        storage_svc = storage or get_storage_service()
        normalized = StorageService.normalize_path(file_path)
        if storage_svc.exists(normalized):
            content = storage_svc.get(normalized)
        else:
            path = Path(file_path)
            if storage_root and not path.is_absolute():
                path = UploadService.resolve_storage_path(file_path, storage_root)
            if not path.exists():
                return []
            content = path.read_bytes()
        if mime_type == "application/json" or filename.lower().endswith(".json"):
            return UploadService.read_json_edc_rows(
                content, dataset_name, max_rows=max_rows
            )
        sheets = UploadService._parse_file(content, mime_type, filename)
        target = next(
            (s for s in sheets if s["name"] == dataset_name),
            sheets[0] if sheets else None,
        )
        if target is None:
            return []

        if mime_type == "text/csv" or filename.lower().endswith(".csv"):
            text = content.decode("utf-8", errors="replace")
            reader = csv.reader(io.StringIO(text))
            all_rows = list(reader)
            if not all_rows:
                return []
            headers = all_rows[0]
            data_rows = all_rows[1 : max_rows + 1]
            rows_out = []
            for row in data_rows:
                record = {}
                for idx, header in enumerate(headers):
                    key = header or f"col_{idx}"
                    record[key] = row[idx] if idx < len(row) else None
                rows_out.append(record)
            return rows_out

        import openpyxl

        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True, read_only=True)
        sheet_names = wb.sheetnames
        if dataset_name in sheet_names:
            ws = wb[dataset_name]
        else:
            sanitized = UploadService._sanitize_sheet_name(dataset_name)
            form_prefix = dataset_name.split(" — ", 1)[0].strip()
            matched = next(
                (
                    name
                    for name in sheet_names
                    if name == sanitized
                    or name == form_prefix
                    or name.startswith(f"{form_prefix} ")
                ),
                None,
            )
            ws = wb[matched or sheet_names[0]]
        values = list(ws.values)
        wb.close()
        if not values:
            return []
        headers = [
            str(h) if h is not None else f"col_{i}" for i, h in enumerate(values[0])
        ]
        rows_out = []
        for row in values[1 : max_rows + 1]:
            record = {}
            for idx, header in enumerate(headers):
                val = row[idx] if idx < len(row) else None
                record[header] = str(val) if val is not None else None
            rows_out.append(record)
        return rows_out

    @staticmethod
    def _profile_column(column_name: str, values: list[str | None]) -> dict:
        """Compute profiling stats for a single column's raw string values."""
        non_null = [v for v in values if v is not None and v != ""]
        sample = mask_sample_values(column_name, non_null[:5])
        missing = sum(1 for v in values if v is None or v == "")
        distinct = len(set(non_null))

        inferred_type = "string"
        probe = non_null[:50]
        if probe:
            if all(_try_float(v) is not None for v in probe):
                inferred_type = "number"
            elif all(_try_date(v) for v in probe):
                inferred_type = "date"
            elif all(
                v.strip().lower() in ("true", "false", "yes", "no", "1", "0")
                for v in probe
            ):
                inferred_type = "boolean"

        min_val = max_val = None
        if inferred_type == "number" and non_null:
            try:
                nums = [float(v) for v in non_null]
                min_val = str(min(nums))
                max_val = str(max(nums))
            except Exception:
                pass

        return {
            "sample_values": sample,
            "missing_count": missing,
            "distinct_count": distinct,
            "inferred_type": inferred_type,
            "min_value": min_val,
            "max_value": max_val,
        }

    async def _count_synthetic_uploads(
        self, study_id: UUID, organization_id: UUID
    ) -> int:
        """Count prior synthetic uploads for version labeling."""
        items, _ = await self._repo.list_for_study(
            study_id, organization_id, limit=1000, offset=0
        )
        return sum(
            1
            for upload in items
            if upload.data_source_type == DataSourceType.SYNTHETIC
            or upload.is_synthetic
        )

    @staticmethod
    def _detect_synthetic_upload(content: bytes, mime_type: str, filename: str) -> bool:
        """Infer synthetic provenance from export filename or embedded markers."""
        if "synthetic" in filename.lower():
            return True
        if mime_type in _PARSEABLE_MIME_TYPES:
            try:
                sample = content[:8192].decode("utf-8", errors="replace")
                if "SYNTHETIC" in sample:
                    return True
                lowered = sample.lower()
                if "synthetic" in lowered and (
                    "label" in lowered or "document_type" in lowered
                ):
                    return True
            except Exception:
                pass
        return False

    @staticmethod
    def _extract_legacy_metadata(content: bytes, mime_type: str) -> dict:
        """Backward-compatible metadata for the extracted_metadata column."""
        if mime_type == "text/csv":
            try:
                text = content.decode("utf-8", errors="replace")
                reader = csv.reader(io.StringIO(text))
                rows = list(reader)
                if rows:
                    return {
                        "columns": rows[0],
                        "row_count": len(rows) - 1,
                        "format": "csv",
                    }
            except Exception:
                pass
        return {"format": mime_type}


def _try_float(v: str) -> float | None:
    try:
        return float(v.replace(",", ""))
    except (ValueError, AttributeError):
        return None


def _try_date(v: str) -> bool:
    from datetime import datetime as _dt

    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%Y%m%d"):
        try:
            _dt.strptime(v.strip(), fmt)
            return True
        except (ValueError, AttributeError):
            pass
    return False
